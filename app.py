import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====================== CẤU HÌNH CHI TIẾT 6 MẪU NGHIỆP VỤ ======================
MAU_MASTER_CONFIG = {
    "Mau 01": {
        "title": "KHOA - PHÒNG - GIƯỜNG",
        "theme": "#1E40AF", "icon": "🏥",
        "desc": "Cột chính: STT, MA_KHOA, TEN_KHOA, BAN_KHAM, GIUONG_TK...",
        "sheet": "MAU_01",
        "out_cols": ["STT", "MA_KHOA", "TEN_KHOA", "BAN_KHAM", "GIUONG_PD", "GIUONG_TK", "GIUONG_HSTC", "GIUONG_HSCC", "TU_NGAY", "DEN_NGAY", "MA_CSKCB"],
        "mapping": {}, "need_input": True, "fixed_val": {"TU_NGAY": "20260101"}
    },
    "Mau 02": {
        "title": "DANH MỤC NHÂN VIÊN Y TẾ",
        "theme": "#0369A1", "icon": "👨‍⚕️",
        "desc": "Cột chính: HO_TEN, SO_CCCD, MACCHN, Phạm vi chuyên môn...",
        "sheet": "MAU_02",
        "out_cols": ["STT", "MA_KHOA", "TEN_KHOA", "HO_TEN", "GIOI_TINH", "SO_DINH_DANH", "CHUCDANH_NN", "VI_TRI", "MACCHN", "NGAYCAP_CCHN", "NOICAP_CCHN", "PHAMVI_CM", "PHAMVI_CMBS", "DVKT_KHAC", "VB_PHANCONG", "THOIGIAN_DK", "THOIGIAN_NGAY", "THOIGIAN_TUAN", "CSKCB_KHAC", "CSKCB_CGKT", "QD_CGKT", "TU_NGAY", "DEN_NGAY", "MA_CSKCB"],
        "mapping": {"SO_CCCD": "SO_DINH_DANH"}, "need_input": True
    },
    "Mau 03": {
        "title": "DANH MỤC THUỐC",
        "theme": "#059669", "icon": "💊",
        "desc": "Cột chính: MA_THUOC, TEN_HOAT_CHAT, SO_DANG_KY, DON_GIA...",
        "sheet": "MAU_03",
        "out_cols": ["STT", "MA_THUOC", "TEN_HOAT_CHAT", "TEN_THUOC", "DON_VI_TINH", "HAM_LUONG", "DUONG_DUNG", "DANG_BAO_CHE", "SO_DANG_KY", "SO_LUONG", "DON_GIA", "DON_GIA_BH", "QUY_CACH", "NHA_SX", "NUOC_SX", "NHA_THAU", "TT_THAU", "TU_NGAY", "DEN_NGAY", "MA_CSKCB", "LOAI_THUOC", "LOAI_THAU", "HT_THAU"],
        "mapping": {}, "need_input": False
    },
    "Mau 04": {
        "title": "DANH MỤC VẬT TƯ Y TẾ",
        "theme": "#D97706", "icon": "📦",
        "desc": "Cột chính: MA_VAT_TU, TEN_VAT_TU, DON_GIA, NHA_THAU...",
        "sheet": "MAU_04",
        "out_cols": ["STT", "MA_VAT_TU", "NHOM_VAT_TU", "TEN_VAT_TU", "MA_HIEU", "SO_LUU_HANH", "TINHNANG_KT", "QUY_CACH", "HANG_SX", "NUOC_SX", "DON_VI_TINH", "DON_GIA", "DON_GIA_BH", "TYLE_TT_BH", "SO_LUONG", "DINH_MUC", "NHA_THAU", "TT_THAU", "TU_NGAY_HD", "DEN_NGAY_HD", "MA_CSKCB", "LOAI_THAU", "HT_THAU", "MA_CSKCB_TBYT", "TU_NGAY", "DEN_NGAY"],
        "mapping": {"TU_NGAY": "TU_NGAY_HD"}, "need_input": False
    },
    "Mau 05": {
        "title": "DỊCH VỤ KỸ THUẬT",
        "theme": "#7C3AED", "icon": "🧪",
        "desc": "Cột chính: MA_TUONG_DUONG, TEN_DVKT, PHAN_LOAI_PTTT...",
        "sheet": "MAU_05",
        "out_cols": ["STT", "MA_DICH_VU", "TEN_DICH_VU", "TEN_DVKT_GIA", "DON_GIA", "QUY_TRINH", "SO_LUONG_CGKT", "CSKCB_CGKT", "CSKCB_CLS", "QD_DVKT", "QD_PD_GIA", "GHI_CHU", "TU_NGAY", "DEN_NGAY", "MA_CSKCB", "GIA_THAN_TOAN"],
        "mapping": {"MA_TUONG_DUONG": "MA_DICH_VU", "TEN_DVKT_PHEDUYET": "TEN_DICH_VU", "TUNGAY": "TU_NGAY", "DENNGAY": "DEN_NGAY"}, "need_input": True
    },
    "Mau 06": {
        "title": "TRANG THIẾT BỊ Y TẾ",
        "theme": "#475569", "icon": "⚙️",
        "desc": "Cột chính: TEN_TB, KY_HIEU, NAM_SX, MA_MAY, SO_LUU_HANH...",
        "sheet": "MAU_06",
        "out_cols": ["STT", "TEN_TB", "KY_HIEU", "CONGTY_SX", "NUOC_SX", "NAM_SX", "NAM_SD", "MA_MAY", "SO_LUU_HANH", "HD_TU", "HD_DEN", "TU_NGAY", "DEN_NGAY", "MA_CSKCB"],
        "mapping": {}, "need_input": True
    }
}

class DataExportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hệ Thống Kết Xuất Dữ Liệu BHYT")
        self.root.geometry("550x720")
        self.root.configure(bg="#F8FAFC")
        self.file_path = ""

        # --- HEADER ---
        header = tk.Frame(self.root, bg="#1B2A4A", height=80)
        header.pack(fill="x")
        tk.Label(header, text="XUẤT DỮ LIỆU", font=("Segoe UI", 18, "bold"), fg="white", bg="#1B2A4A").place(relx=0.5, rely=0.5, anchor="center")

        container = tk.Frame(self.root, bg="white", bd=1, relief="solid")
        container.pack(padx=30, pady=20, fill="both", expand=True)

        # 1. Chọn mẫu
        tk.Label(container, text="🔹 CHỌN MẪU BÁO CÁO:", font=("Segoe UI", 10, "bold"), bg="white").pack(anchor="w", padx=25, pady=(20, 5))
        self.cbo_mau = ttk.Combobox(container, values=list(MAU_MASTER_CONFIG.keys()), font=("Segoe UI", 11), state="readonly")
        self.cbo_mau.pack(fill="x", padx=25, pady=5)
        self.cbo_mau.set("Mau 01")
        self.cbo_mau.bind("<<ComboboxSelected>>", self.update_template_ui)

        # Thông tin mẫu
        self.info_box = tk.Frame(container, bg="#F1F5F9")
        self.info_box.pack(fill="x", padx=25, pady=15)
        self.lbl_info_icon = tk.Label(self.info_box, text="🏥", font=("Segoe UI", 24), bg="#F1F5F9")
        self.lbl_info_icon.pack(side="left", padx=15)
        
        txt_frame = tk.Frame(self.info_box, bg="#F1F5F9")
        txt_frame.pack(side="left", fill="both", expand=True, pady=10)
        self.lbl_info_title = tk.Label(txt_frame, text="KHOA - PHÒNG - GIƯỜNG", font=("Segoe UI", 11, "bold"), fg="#1E293B", bg="#F1F5F9", anchor="w")
        self.lbl_info_title.pack(fill="x")
        self.lbl_info_desc = tk.Label(txt_frame, text="Yêu cầu file đầu vào...", font=("Segoe UI", 9), fg="#64748B", bg="#F1F5F9", anchor="w", justify="left", wraplength=300)
        self.lbl_info_desc.pack(fill="x")

        # 2. MACSKCB & MA_CSKCB
        tk.Label(container, text="Mã đơn vị (MACSKCB):", font=("Segoe UI", 9, "bold"), bg="white").pack(anchor="w", padx=25, pady=(5, 0))
        self.ent_ma_dv = tk.Entry(container, font=("Segoe UI", 11), bd=1, relief="solid")
        self.ent_ma_dv.pack(fill="x", padx=25, pady=5, ipady=3)
        self.ent_ma_dv.insert(0, "MACSKCB")

        tk.Label(container, text="Mã cơ sở (MA_CSKCB):", font=("Segoe UI", 9, "bold"), bg="white").pack(anchor="w", padx=25, pady=(5, 0))
        self.ent_ma_cs = tk.Entry(container, font=("Segoe UI", 11), bd=1, relief="solid")
        self.ent_ma_cs.pack(fill="x", padx=25, pady=5, ipady=3)
        self.ent_ma_cs.insert(0, "84003")

        # 3. Chọn file (DÒNG 121 ĐÃ SỬA TẠI ĐÂY)
        tk.Label(container, text="📂 FILE DỮ LIỆU ĐẦU VÀO:", font=("Segoe UI", 9, "bold"), bg="white").pack(anchor="w", padx=25, pady=(15, 0))
        f_frame = tk.Frame(container, bg="white")
        f_frame.pack(fill="x", padx=25, pady=5)
        
        # Đã đổi relief="dashed" thành relief="solid"
        self.lbl_file = tk.Label(f_frame, text="Chưa chọn file...", font=("Segoe UI", 10), bg="#F8FAFC", bd=1, relief="solid", anchor="w")
        self.lbl_file.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 5))
        tk.Button(f_frame, text="Chọn", bg="#64748B", fg="white", font=("Segoe UI", 9, "bold"), command=self.browse_file).pack(side="right")

        # 4. Nút xuất
        self.btn_export = tk.Button(container, text="🚀 BẮT ĐẦU XUẤT EXCEL", font=("Segoe UI", 12, "bold"), bg="#10B981", fg="white", command=self.process_data, cursor="hand2", bd=0, pady=12)
        self.btn_export.pack(fill="x", padx=25, pady=30)

        self.update_template_ui()

    def update_template_ui(self, event=None):
        mau = self.cbo_mau.get()
        cfg = MAU_MASTER_CONFIG[mau]
        self.lbl_info_icon.config(text=cfg["icon"])
        self.lbl_info_title.config(text=cfg["title"], fg=cfg["theme"])
        self.lbl_info_desc.config(text=cfg["desc"])
        self.btn_export.config(bg=cfg["theme"])
        state = "normal" if cfg["need_input"] else "disabled"
        bg_col = "white" if cfg["need_input"] else "#F1F5F9"
        self.ent_ma_dv.config(state=state, bg=bg_col)
        self.ent_ma_cs.config(state=state, bg=bg_col)

    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.file_path = path
            self.lbl_file.config(text=os.path.basename(path), fg="#1E40AF")

    def process_data(self):
        if not self.file_path:
            messagebox.showwarning("Lỗi", "Vui lòng chọn file!")
            return
        mau_sel = self.cbo_mau.get()
        cfg = MAU_MASTER_CONFIG[mau_sel]
        ma_dv = self.ent_ma_dv.get().strip()
        ma_cs = self.ent_ma_cs.get().strip()
        try:
            df_in = pd.read_excel(self.file_path, dtype=str)
            df_out = pd.DataFrame(columns=cfg["out_cols"])
            for col_out in cfg["out_cols"]:
                col_in = next((k for k, v in cfg["mapping"].items() if v == col_out), col_out)
                df_out[col_out] = df_in[col_in] if col_in in df_in.columns else ""
            if "MA_CSKCB" in df_out.columns: df_out["MA_CSKCB"] = ma_cs
            if "MACSKCB" in df_out.columns: df_out["MACSKCB"] = ma_dv
            if "fixed_val" in cfg:
                for col, val in cfg["fixed_val"].items():
                    if col in df_out.columns: df_out[col] = val
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"Daura_{cfg['sheet']}.xlsx")
            if save_path:
                df_out.to_excel(save_path, index=False, sheet_name=cfg["sheet"])
                self._format_excel(save_path, cfg["theme"].replace("#", ""))
                messagebox.showinfo("Thành công", f"Đã xuất {cfg['title']} thành công!")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def _format_excel(self, path, theme_color):
        wb = load_workbook(path)
        ws = wb.active
        h_fill = PatternFill("solid", fgColor=theme_color)
        h_font = Font(name="Calibri", bold=True, color="FFFFFF")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill, cell.font, cell.border = h_fill, h_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col] + [10])
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        wb.save(path)

if __name__ == "__main__":
    root = tk.Tk()
    app = DataExportApp(root)
    root.mainloop()